"""
BOM Checker
"""
import sys
import argparse
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Keywords used to auto-detect columns (case-insensitive, partial match)
MFR_KEYWORDS = ["manufacturer", "mfr", "mfg"]
LIFECYCLE_KEYWORDS = ["lifecycle", "life cycle", "status", "life-cycle"]

# Values that indicate a part is End-of-Life / obsolete
EOL_KEYWORDS = [
    "eol",
    "end of life",
    "obsolete",
    "discontinued",
    "not recommended for new designs",
    "nrnd",
    "last time buy",
    "ltb",
]


def find_column(columns, keywords):
    """Return the first column name whose lowercase text contains any keyword."""
    for col in columns:
        col_lower = str(col).lower()
        for kw in keywords:
            if kw in col_lower:
                return col
    return None


def is_blank(value):
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def is_eol(value):
    if is_blank(value):
        return False
    val_lower = str(value).lower()
    return any(kw in val_lower for kw in EOL_KEYWORDS)


def analyze_bom(df):
    """Run checks and return the dataframe with an added 'Issues' column,
    plus the detected column names."""
    mfr_col = find_column(df.columns, MFR_KEYWORDS)
    lifecycle_col = find_column(df.columns, LIFECYCLE_KEYWORDS)

    issues_list = []
    for _, row in df.iterrows():
        row_issues = []

        if mfr_col:
            if is_blank(row[mfr_col]):
                row_issues.append("Missing Manufacturer")
        else:
            row_issues.append("Manufacturer column not found")

        if lifecycle_col:
            lifecycle = str(row[lifecycle_col]).upper()

            if is_blank(row[lifecycle_col]):
                row_issues.append("Missing Lifecycle Status")

            elif lifecycle == "OBSOLETE":
                row_issues.append("Obsolete Part")

            elif lifecycle == "LTB":
                row_issues.append("LTB Part")

            elif lifecycle == "NRND":
                row_issues.append("NRND Part")
        else:
            row_issues.append("Lifecycle Status column not found")

        issues_list.append("; ".join(row_issues) if row_issues else "")

    df = df.copy()
    df["Issues"] = issues_list
    return df, mfr_col, lifecycle_col

def classify_part(description):
    """Classify BOM row into component type."""
    if pd.isna(description):
        return "Other"

    desc = str(description).upper()

    if any(x in desc for x in ["CONN", "CONNECTOR", "HEADER", "USB", "JACK", "SOCKET"]):
        return "Connector"

    if any(x in desc for x in ["RES ", "RESISTOR", "OHM"]):
        return "Resistor"

    if any(x in desc for x in ["CAP ", "CAPACITOR", "UF", "NF", "PF"]):
        return "Capacitor"

    if any(x in desc for x in ["IND ", "INDUCTOR", "CHOKE"]):
        return "Inductor"

    if any(x in desc for x in ["BATTERY", "CELL", "LIPO", "LI-ION"]):
        return "Battery"

    if any(x in desc for x in ["CRYSTAL", "XTAL", "OSCILLATOR"]):
        return "Crystal"

    if any(x in desc for x in ["IC", "MCU", "CPU", "FPGA", "ASIC", "SENSOR"]):
        return "IC"

    return "Other"


def add_part_type_column(df):
    """Add Part_Type column using DESCRIPTION."""
    if "DESCRIPTION" in df.columns:
        df["Part_Type"] = df["DESCRIPTION"].apply(classify_part)
    else:
        df["Part_Type"] = "Other"

    return df

LIFECYCLE_DB = {
    "CRCW02011K00FKED": "Active",
    "ERJ-1GNF1002C-ND": "NRND",
    "54548-2272": "LTB",
    "DX07S024JJ7": "Obsolete"
}

def get_lifecycle_status(mpn):
    return LIFECYCLE_DB.get(str(mpn).strip(), "Unknown")

def get_risk_level(lifecycle):
    lifecycle = lifecycle.upper()

    if lifecycle == "ACTIVE":
        return "No Risk"
    elif lifecycle == "NRND":
        return "Medium Risk"
    elif lifecycle in ["LTB", "OBSOLETE"]:
        return "High Risk"
    else:
        return "Review Required"

def add_lifecycle_risk_columns(df):
    df["Lifecycle_Status"] = df["MPN"].apply(get_lifecycle_status)
    df["Risk_Level"] = df["Lifecycle_Status"].apply(get_risk_level)

    df["Alternate_Required"] = df["Risk_Level"].apply(
        lambda x: "Yes" if x in ["Medium Risk", "High Risk"] else "No"
    )

    return df
def write_report(df, output_path, sheet_name="BOM"):
    df.to_excel(output_path, index=False, sheet_name=sheet_name)

    # Highlight rows with issues
    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    issues_col_idx = header.index("Issues") + 1

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        issue_cell = row[issues_col_idx - 1]
        issue_text = str(issue_cell.value) if issue_cell.value else ""
        if "Obsolete Part" in issue_text:
            fill = red_fill

        elif "LTB Part" in issue_text:
            fill = red_fill

        elif "NRND Part" in issue_text:
            fill = yellow_fill

        elif issue_text:
            fill = yellow_fill
        else:
            fill = None
        if fill:
            for cell in row:
                cell.fill = fill

    wb.save(output_path)


def print_summary(df, mfr_col, lifecycle_col):
    total = len(df)
    missing_mfr = df["Issues"].str.contains("Missing Manufacturer").sum() if mfr_col else "N/A (column not found)"
    missing_lifecycle = df["Issues"].str.contains("Missing Lifecycle Status").sum() if lifecycle_col else "N/A (column not found)"
    eol_parts = df["Issues"].str.contains("EOL Part").sum()
    total_issues = len(df[df["Issues"] != ""])

    print("\n===== BOM Check Summary =====")
    print(f"Total line items:            {total}")
    print(f"Manufacturer column:         {mfr_col or 'NOT FOUND'}")
    print(f"Lifecycle Status column:     {lifecycle_col or 'NOT FOUND'}")
    print(f"Missing Manufacturer:        {missing_mfr}")
    print(f"Missing Lifecycle Status:    {missing_lifecycle}")
    print(f"EOL Parts:                   {eol_parts}")
    print(f"Rows with any issue:         {total_issues}")
    print("==============================\n")


def main():
    parser = argparse.ArgumentParser(
        description="Check a BOM Excel file for missing manufacturer, missing lifecycle status, and EOL parts."
    )

    parser.add_argument(
        "bom_file",
        help="Path to the input BOM Excel file (.xlsx)"
    )

    parser.add_argument(
        "--sheet",
        default=0,
        help="Sheet name or index to read (default: first sheet)"
    )

    parser.add_argument(
        "--out",
        default=None,
        help="Path to save the annotated report"
    )

    args = parser.parse_args()

    try:
        df = pd.read_excel(args.bom_file, sheet_name=args.sheet)

        # Add component classification
        df = add_part_type_column(df)

        # Add lifecycle and risk columns
        df = add_lifecycle_risk_columns(df)

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        sys.exit(1)

    df, mfr_col, lifecycle_col = analyze_bom(df)

    from datetime import datetime

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_path = args.out or args.bom_file.replace(
        ".xlsx",
        f"_checked_{timestamp}.xlsx"
    )

    write_report(df, output_path)

    print_summary(df, mfr_col, lifecycle_col)
    print(f"Annotated report saved to: {output_path}")


if __name__ == "__main__":
    main()